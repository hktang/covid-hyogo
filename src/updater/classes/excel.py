from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook
import urllib.request


class Excel:
    """Excel document class"""

    params = {
        'url': 'https://web.pref.hyogo.lg.jp/kk03/documents/corona-kanjajokyou.xlsx',
        'sheet_name': '1',
        'start_row': 6,
        'date_cols': ['C', 'I'],
        'last_updated_cell': 'Y1',
        'annotation_cell': 'I3',
        'time_format': "%Y-%m-%d",
    }

    workbook = None
    worksheet = None

    def __init__(self, params=params):
        self.params.update(params)
        self.workbook = self.load_workbook_from_url()

    def load_workbook_from_url(self):
        ''' Loads workbook based on URL. '''
        file = urllib.request.urlopen(self.params['url']).read()
        return load_workbook(filename=BytesIO(file), data_only=True)

    def load_worksheet(self):
        ''' Loads sheet based on sheet name. '''
        return self.workbook[self.params['sheet_name']]

    def ordinal_to_date(self, ordinal, _epoch0=datetime(1899, 12, 31)):
        ''' 
            Convert integer to formatted date.
            Courtesy of https://stackoverflow.com/a/29387450/1017265
        '''
        if ordinal >= 60:
            ordinal -= 1
        return (_epoch0 + timedelta(days=ordinal)).replace(microsecond=0)

    def get_column_letter(self, n):
        ''' Returns column letter A, AA, etc., based on col number.'''
        letter = ''
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter

    def get_case_data(self, ws):
        ''' Returns cleaned data of all cases.'''
        case_data = []
        count = 0
        max_row = ws.max_row

        for row in ws:
            count += 1
            row_data = []

            # Skip header rows and trailing empty rows
            if (count < self.params['start_row'] or count > max_row):
                continue

            if row[1].value == None:
                continue

            for cell in row:

                # Convert None to empty string
                if cell.value == None:
                    row_data.append('')
                    continue

                if cell.column_letter in self.params['date_cols']:
                    if isinstance(cell.value, int):
                        row_data.append(self.ordinal_to_date(
                            cell.value).strftime(self.params['time_format']))
                    elif isinstance(cell.value, datetime):
                        row_data.append(cell.value.strftime(
                            self.params['time_format']))
                    else:
                        row_data.append(str(cell.value))
                else:
                    row_data.append(str(cell.value))

            case_data.append(row_data)

        return case_data

    def get_newest_date(self, ws):
        ''' Returns the formatted date of the newest date of cases'''
        date_value = ws[self.params['date_cols']
                        [0]+str(self.params['start_row'])].value
        return self.ordinal_to_date(date_value)
